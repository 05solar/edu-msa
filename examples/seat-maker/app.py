import io
import os
import random
from typing import List, Optional
from urllib.parse import quote

from fastapi import FastAPI, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, FileResponse
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

from ui import INDEX

app = FastAPI(title="seat-maker")


def err(code: str, message: str, status: int = 400):
    return JSONResponse({"error": {"code": code, "message": message}}, status_code=status)


class Student(BaseModel):
    name: str
    gender: Optional[str] = None  # M | F | None


class Fixed(BaseModel):
    name: str
    r: int
    c: int


class ArrangeReq(BaseModel):
    students: List[Student]
    rows: int
    cols: int
    method: str = "random"          # random | gender | order
    separate: List[List[str]] = []  # [[a,b], ...] 인접 금지
    fixed: List[Fixed] = []
    seed: Optional[int] = None


class ExportReq(BaseModel):
    title: str = "자리 배치표"
    grid: List[List[Optional[str]]]  # 이름 또는 null


NEI = [(-1, 0), (1, 0), (0, -1), (0, 1)]


def _pos(grid):
    p = {}
    for r, row in enumerate(grid):
        for c, cell in enumerate(row):
            if cell:
                p[cell["name"]] = (r, c)
    return p


def _conflicts(grid, separate):
    pos = _pos(grid)
    bad = []
    for pair in separate:
        if len(pair) != 2:
            continue
        a, b = pair[0], pair[1]
        if a in pos and b in pos:
            (ra, ca), (rb, cb) = pos[a], pos[b]
            if abs(ra - rb) + abs(ca - cb) == 1:
                bad.append((a, b))
    return bad


def arrange(req: ArrangeReq):
    rng = random.Random(req.seed)
    rows, cols = req.rows, req.cols
    total = rows * cols
    grid = [[None] * cols for _ in range(rows)]
    by_name = {s.name: {"name": s.name, "gender": s.gender} for s in req.students}
    placed = set()

    # 고정석 우선
    for f in req.fixed:
        if 0 <= f.r < rows and 0 <= f.c < cols and grid[f.r][f.c] is None and f.name in by_name:
            grid[f.r][f.c] = by_name[f.name]
            placed.add(f.name)

    remaining = [by_name[s.name] for s in req.students if s.name not in placed]
    empty = [(r, c) for r in range(rows) for c in range(cols) if grid[r][c] is None]

    if req.method == "order":
        order = remaining
        for (r, c), st in zip(empty, order):
            grid[r][c] = st
    elif req.method == "gender":
        males = [s for s in remaining if s["gender"] == "M"]
        females = [s for s in remaining if s["gender"] == "F"]
        others = [s for s in remaining if s["gender"] not in ("M", "F")]
        rng.shuffle(males)
        rng.shuffle(females)
        rng.shuffle(others)
        even = [(r, c) for (r, c) in empty if (r + c) % 2 == 0]
        odd = [(r, c) for (r, c) in empty if (r + c) % 2 == 1]
        # 남=짝수칸, 여=홀수칸 우선 배치, 넘치면 남는 칸으로
        pool_map = [(even, males), (odd, females)]
        leftover_seats = []
        leftover_people = []
        for seats, people in pool_map:
            n = min(len(seats), len(people))
            for i in range(n):
                grid[seats[i][0]][seats[i][1]] = people[i]
            leftover_seats += seats[n:]
            leftover_people += people[n:]
        leftover_people += others
        for (r, c), st in zip(leftover_seats, leftover_people):
            grid[r][c] = st
    else:  # random
        order = remaining[:]
        rng.shuffle(order)
        for (r, c), st in zip(empty, order):
            grid[r][c] = st

    # 분리(인접 금지) 로컬 스왑 보정
    satisfied = True
    if req.separate:
        cells = [(r, c) for r in range(rows) for c in range(cols)]
        for _ in range(400):
            bad = _conflicts(grid, req.separate)
            if not bad:
                break
            pos = _pos(grid)
            a = bad[0][0]
            ra, ca = pos[a]
            r2, c2 = rng.choice(cells)
            grid[ra][ca], grid[r2][c2] = grid[r2][c2], grid[ra][ca]
        satisfied = len(_conflicts(grid, req.separate)) == 0

    unplaced = [s["name"] for s in remaining[max(0, len(empty)):]] if len(remaining) > len(empty) else []
    return grid, unplaced, satisfied


@app.get("/healthz")
def healthz():
    return {"status": "ok", "service": "seat-maker"}


@app.post("/api/arrange")
def api_arrange(req: ArrangeReq):
    if not req.students:
        return err("VALIDATION", "학생 명단을 입력하세요.")
    if req.rows < 1 or req.cols < 1 or req.rows > 20 or req.cols > 20:
        return err("VALIDATION", "행/열은 1~20 사이여야 합니다.")
    if len(req.students) > req.rows * req.cols:
        return err("VALIDATION", f"좌석({req.rows}x{req.cols}={req.rows*req.cols})보다 학생 수({len(req.students)})가 많습니다.")
    grid, unplaced, satisfied = arrange(req)
    return {
        "grid": grid,
        "rows": req.rows,
        "cols": req.cols,
        "unplaced": unplaced,
        "separateSatisfied": satisfied,
        "count": len(req.students),
    }


@app.post("/api/parse-upload")
async def parse_upload(file: UploadFile = File(...)):
    data = await file.read()
    name = (file.filename or "").lower()
    students = []
    try:
        if name.endswith(".xlsx") or name.endswith(".xlsm"):
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                nm = str(row[0]).strip()
                if not nm or nm in ("이름", "성명", "name"):
                    continue
                g = None
                if len(row) > 1 and row[1] is not None:
                    gv = str(row[1]).strip()
                    if gv in ("남", "M", "m", "male", "남자"):
                        g = "M"
                    elif gv in ("여", "F", "f", "female", "여자"):
                        g = "F"
                students.append({"name": nm, "gender": g})
        else:  # csv / txt
            text = data.decode("utf-8-sig", errors="replace")
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                parts = [p.strip() for p in line.replace("\t", ",").split(",")]
                nm = parts[0]
                if not nm or nm in ("이름", "성명", "name"):
                    continue
                g = None
                if len(parts) > 1:
                    gv = parts[1]
                    if gv in ("남", "M", "m", "male", "남자"):
                        g = "M"
                    elif gv in ("여", "F", "f", "female", "여자"):
                        g = "F"
                students.append({"name": nm, "gender": g})
    except Exception as e:
        return err("PARSE", f"파일을 해석할 수 없습니다: {e}")
    if not students:
        return err("VALIDATION", "명단을 찾을 수 없습니다. 첫 열에 이름, (선택) 둘째 열에 성별을 넣어주세요.")
    return {"students": students}


@app.post("/api/export")
def export_xlsx(req: ExportReq):
    wb = Workbook()
    ws = wb.active
    ws.title = "자리배치"
    thin = Side(style="thin", color="B0B7C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cols = max((len(r) for r in req.grid), default=1)

    ws.cell(row=1, column=1, value=req.title).font = Font(bold=True, size=14)
    tstart = 3
    # 교탁 표시
    tcell = ws.cell(row=tstart, column=1, value="교 탁 (칠판)")
    tcell.alignment = center
    tcell.fill = PatternFill("solid", fgColor="E8EDF5")
    tcell.font = Font(bold=True)
    if cols > 1:
        ws.merge_cells(start_row=tstart, start_column=1, end_row=tstart, end_column=cols)

    for ri, row in enumerate(req.grid):
        for ci in range(cols):
            v = row[ci] if ci < len(row) else None
            cell = ws.cell(row=tstart + 2 + ri, column=ci + 1, value=(v or ""))
            cell.alignment = center
            cell.border = border
            if not v:
                cell.fill = PatternFill("solid", fgColor="F5F6F8")
    for ci in range(1, cols + 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 12
    for ri in range(len(req.grid)):
        ws.row_dimensions[tstart + 2 + ri].height = 34

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = quote("자리배치표.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )


@app.get("/og.png")
def og():
    return FileResponse(os.path.join(os.path.dirname(__file__), "og.png"), media_type="image/png")


@app.get("/", response_class=HTMLResponse)
def index():
    return INDEX
